print("Starting")
file1 = open("FORMAT.py","w") # the name of the output file, the one that will contend the format code

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import re

rb = load_workbook(filename = 'formato.xlsx') # name of your input file # Read Book

rs=rb.active # Read Sheet

file1.write("#PASTE THE FOLLOWING CODE IN A .py FILE")
file1.write("\n")
file1.write("\n")
file1.write("\n")
file1.write("from openpyxl import Workbook")
file1.write("\n")
file1.write("from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font")
file1.write("\n")

file1.write("wb = Workbook()") #Write Book # the book that you will format
file1.write("\n")
file1.write("ws=wb.active") # Write Sheet
file1.write("\n")
file1.write("\n")
file1.write("\n")
STARTING = str("STARTING")
WORKING = str("WORKING")
READY = str("THE FILE IS READY")
file1.write("print(\""+STARTING+"\")")
file1.write("\n")

print("Extracting Merged cells")# MERGE WORKS PERFECTLY #1

MERGE = (rs.merged_cells.ranges)
if len(MERGE) >0:
    MERGE = str(MERGE)
    searchM = re.search("<MergedCellRange ", MERGE) #  busca <CellRange :
    inicioM=searchM.end()
    lenM=len(MERGE)
    MERGE =MERGE[inicioM:lenM]
    searchMM = re.search(">", MERGE)
    MERGE1 =(MERGE[0:searchMM.end()-1])



    MERGE2 = str("\""+MERGE1+"\"")
    MERGE2 = (str("ws.merge_cells")+"("+MERGE2)+")"
    file1.write (MERGE2)
    file1.write("\n")



    while lenM  > 4:

        try:
            searchM = re.search("<MergedCellRange ", MERGE) #  busca <CellRange :
            inicioM=searchM.end()
            lenM=len(MERGE)
            MERGE =MERGE[inicioM:lenM]
            searchMM = re.search(">", MERGE)
            MERGE1 =(MERGE[0:searchMM.end()-1])
            MERGE2 = str("\""+MERGE1+"\"")
            MERGE2 = (str("ws.merge_cells")+"("+MERGE2)+")"
            file1.write(MERGE2)
            file1.write("\n")

        except:
            break


#number formats # DOSENT SEEM TO CHANGE THE FILE IN ANY WAY #2
print(" Extracting Number formats")
file1.write("\n")
file1.write("\n")
file1.write("\n")
for row in rs.rows:
    for cell in row:

        text=(str(cell))#<Cell 'Sheet1'.D8>=None

        #Conviernte <Cell 'Sheet1'.D8>=None en D8
        start =text.find(".")+1
        end = text.find(">", start)
        CELL = text[start:end]
        CELL1 =("\'"+CELL+ "\'")
        #rs['A2'].number_format


        NUMF = ("ws["+CELL1+"].number_format ")#bien
        NUMF0= (rs[CELL].number_format)
        NUMF1=("\'"+NUMF0+ "\'")
        NUMF2 = NUMF+" = " + NUMF1
        if NUMF0 != "General":
            file1.write(NUMF2)
            file1.write("\n")


file1.write("\n")
file1.write("\n")



print("  Extracting Row heights") # WORK JUST FINE

file1.write("#row height:")
for hrow in range (1, rs.max_row+1):
    file1.write("\n")
    HROW1 =  ("hrow" + str(hrow)+ " = " + str(rs.row_dimensions[hrow].height))
    file1.write("\n")
    file1.write(HROW1)
    HROW2 = ("ws.row_dimensions["+str(hrow)+"].height  = hrow" + str(hrow))
    file1.write("\n")
    file1.write(HROW2)
    file1.write("\n")

print("   Extracting Column width") # ITS THE BANE OF MY EXISTENCE, # IT SEEMS TO WORK IN WORKBOOK, BUT NOT IN LOAD_WOORKBOOK IDW

file1.write("#column width:")
for wcol in range (1, rs.max_column):


    file1.write("\n")
    WCOL1= chr(ord('@')+wcol)
    WCOL2 = ("wcol"+str(WCOL1) + " = " +str(rs.column_dimensions[WCOL1].width))
    file1.write("\n")
    file1.write (WCOL2)
    WCOL3 =("ws.column_dimensions[\""+WCOL1+"\"].width = wcol"+str(WCOL1))
    file1.write("\n")
    file1.write (WCOL3)
    file1.write("\n")

print("    Extracting Values")
file1.write("\n")
for row in rs.rows:
    for cell in row:


        #print(cell.value)

        text=(str(cell))#<Cell 'Sheet1'.D8>=None


        #Conviernte <Cell 'Sheet1'.D8>=None en D8
        start =text.find(".")+1
        end = text.find(">", start)
        CELL = text[start:end]

        def safeStr(obj):
            try: return str(obj)
            except UnicodeEncodeError:
                return obj.encode('ascii', 'ignore').decode('ascii')
            except: return ""

        CONT = str("ws[\""+CELL+ "\"] = ")
        CONT1 = safeStr(cell.value)
        #CONT1 = CONT1.decode('ascii', 'ignore')
        #print(str(cell))
        if CONT1 != "None":
            file1.write(CONT + "\""+CONT1 +"\"")
            file1.write("\n")
            #print(CONT + "\""+CONT1 +"\"")


print("     Extracting Fonts")
i = 0
for row in rs.rows:
    for cell in row:
        i = i +1

        #print(cell.value)
        text=(str(cell))#<Cell 'Sheet1'.D8>=None


        #Conviernte <Cell 'Sheet1'.D8>=None en D8
        start =text.find(".")+1
        end = text.find(">", start)
        CELL = text[start:end]

        #Font
        X = ("%s"% (CELL,))
        FONT=rs[X].font
        FONT = str(FONT)
        text=(FONT)

        S1 =("name=")
        start1 =text.find(S1)+len(S1)
        end1 = text.find(",", start1)
        name = text[start1:end1]

        S2 =("sz=")
        start2 =text.find(S2)+len(S2)
        end2 = text.find(",", start2)
        size = text[start2:end2]

        S3 =("b=")
        start3 =text.find(S3)+len(S3)
        end3 = text.find(",", start3)
        bold = text[start3:end3]

        S4 =("i=")
        start4 =text.find(S4)+len(S4)
        end4 = text.find(",", start4)
        italic = text[start4:end4]

        S5 =("vertAlign=")
        start5 =text.find(S5)+len(S5)
        end5 = text.find(",", start5)
        vertAlign = text[start5:end5]

        S6 =("u=")
        start6 =text.find(S6)+len(S6)
        end6 = text.find(",", start6)
        underline = text[start6:end6]

        S7 =("strike=")
        start7 =text.find(S7)+len(S7)
        end7 = text.find(",", start7)
        strike = text[start7:end7]

        S8 =("rgb=")
        start8 =text.find(S8)+len(S8)
        end8 = text.find(",", start8)
        color = text[start8:end8]


        FONT1 = str("font"+CELL + " = Font(name="+name+
        ",size="+size+
        ",bold="+bold+
        ",italic="+italic+
        ",vertAlign="+vertAlign+
        ",underline="+underline+
        ",strike="+strike+
        ",color="+color+
        ")")


        file1.write(FONT1)
        file1.write("\n")

        FONT2 = "ws[\""+CELL+ "\"].font = font"+CELL
        FONT2= str(FONT2)
        file1.write(FONT2)

        file1.write("\n")
        file1.write("\n")



        #Alignment
print("      Extracting Alignments")

for row in rs.rows:
    for cell in row:

        #print(cell.value)
        text=(str(cell))#<Cell 'Sheet1'.D8>=None


        #Conviernte <Cell 'Sheet1'.D8>=None en D8
        start =text.find(".")+1
        end = text.find(">", start)
        CELL = text[start:end]

        X = ("%s"% (CELL,))
        ALIG=rs[X].alignment
        ALIG = str(ALIG)
        text=(ALIG)

        S9 =("horizontal=")
        start9 =text.find(S9)+len(S9)
        end9 = text.find(",", start9)
        horizontal = text[start9:end9]

        S10 =("vertical=")
        start10 =text.find(S10)+len(S10)
        end10 = text.find(",", start10)
        vertical = text[start10:end10]

        S11 =("textRotation=")
        start11 =text.find(S11)+len(S11)
        end11 = text.find(",", start11)
        text_rotation = text[start11:end11]

        S12 =("wrapText=")
        start12 =text.find(S12)+len(S12)
        end12 = text.find(",", start12)
        wrap_text = text[start12:end12]

        S13 =("shrinkToFit=")
        start13 =text.find(S13)+len(S13)
        end13 = text.find(",", start13)
        shrink_to_fit = text[start13:end13]

        S14 =("indent=")
        start14 =text.find(S14)+len(S14)
        end14 = text.find(",", start14)
        indent = text[start14:end14]


        ALIG0= str("Alignment(horizontal="+horizontal+
        ",vertical="+vertical+
        ",text_rotation="+text_rotation+
        ",wrap_text="+wrap_text+
        ",shrink_to_fit="+shrink_to_fit+
        ",indent="+indent+")")

        ALIG1 = str("alig"+CELL + " = "+ALIG0)

        ALIG2 = "ws[\""+CELL+ "\"].alignment = alig"+CELL
        ALIG2= str(ALIG2)
        ALIR = "Alignment(horizontal=None,vertical=None,text_rotation=0,wrap_text=None,shrink_to_fit=None,indent=0.0)"

        if ALIG0 != ALIR:
            file1.write(ALIG1)
            file1.write("\n")
            file1.write(ALIG2)
            file1.write("\n")
            file1.write("\n")



#Fill
print("       Extracting Fills")

for row in rs.rows:
    for cell in row:

        #print(cell.value)
        text=(str(cell))#<Cell 'Sheet1'.D8>=None


        #Conviernte <Cell 'Sheet1'.D8>=None en D8
        start =text.find(".")+1
        end = text.find(">", start)
        CELL = text[start:end]

        X = ("%s"% (CELL,))
        FILL=rs[X].fill
        FILL = str(FILL)
        text=(FILL)

        S15 =("patternType=")
        start15 =text.find(S15)+len(S15)
        end15 = text.find(",", start15)
        fill_type = text[start15:end15]

        S16 =("fgColor=")
        start16 =text.find(S16)+62#len(S16)# Why? because fuck new lines thats why.
        end16 = text.find(",", start16)
        start_color = text[start16:end16]

        S17 =("bgColor=")
        start17 =text.find(S17)+62#len(S17)
        end17 = text.find(",", start17)
        end_color = text[start17:end17]

        FILL0 =str("PatternFill(fill_type="+fill_type+
        ",start_color="+start_color+
        ",end_color="+end_color+
        ")")

        FILL1 = str("fill"+CELL + " = PatternFill(fill_type="+fill_type+
        ",start_color="+start_color+
        ",end_color="+end_color+
        ")")



        FILL2 = "ws[\""+CELL+ "\"].fill = fill"+CELL
        FILL2= str(FILL2)
        FILLR = "PatternFill(fill_type=None,start_color='00000000',end_color='00000000')"

        if FILL0 != FILLR:
            file1.write(FILL1)
            file1.write("\n")
            file1.write(FILL2)
            file1.write("\n")
            file1.write("\n")




#Borders          #DIAGONALS NOT INCLUDED.

print("         Extracting Borders")

for row in rs.rows:
    for cell in row:

        #print(cell.value)
        text=(str(cell))#<Cell 'Sheet1'.D8>=None


        #Conviernte <Cell 'Sheet1'.D8>=None en D8
        start =text.find(".")+1
        end = text.find(">", start)
        CELL = text[start:end]



        X = ("%s"% (CELL,))
        BORD=rs[X].border
        BORD = str(BORD)
        text=(BORD)



        text=(BORD)

        #LEFT
        #Sorry the Bord part of the code is potato.

        search1 = re.search("style=", text) # busca style
        inicioN=search1.end()
        lent=len(text)
        text =text[inicioN:lent]
        search2 = re.search(",", text)
        left =(text[0:search2.end()-1])
        if left == "None":
            colorL = "None"
        else:
            try:
                search2 = re.search("rgb=", text)
                textc= text[search2.end():len(text)]
                search3 = re.search(",",textc)
                colorL =(textc[0:search3.end()-1])
            except:
                colorL ="None"


        #RIGTH
        search4 = re.search("style=", text)

        inicioN=search4.end()
        lent=len(text)
        text =text[inicioN:lent]
        search5 = re.search(",", text)

        right = text[0:search5.end()-1]

        if right == "None":
            colorR = "None"
        else:
            try:
                search5 = re.search("rgb=", text)
                textc= text[search5.end():len(text)]
                search6 = re.search(",",textc)
                colorR=(textc[0:search6.end()-1])
            except:
                colorR ="None"



        #TOP

        search7 = re.search("style=", text)

        inicioN=search7.end()
        lent=len(text)
        text =text[inicioN:lent]
        search8 = re.search(",", text)

        top = text[0:search8.end()-1]

        if top == "None":
            colorT = "None"
        else:
            try:
                search8 = re.search("rgb=", text)
                textc= text[search8.end():len(text)]
                search9 = re.search(",",textc)
                colorT=(textc[0:search9.end()-1])
            except:
                colorT = "None"


        #BOTTOM

        search10 = re.search("style=", text)

        inicioN=search10.end()
        lent=len(text)
        text =text[inicioN:lent]
        search11 = re.search(",", text)

        bottom = text[0:search11.end()-1]

        if bottom == "None":
            colorB = "None"
        else:
            try:
                search11 = re.search("rgb=", text)
                textc= text[search11.end():len(text)]
                search12 = re.search(",",textc)
                colorB=(textc[0:search12.end()-1])
            except:
                colorB ="None"

        # end of potato code.
        BORD0= ("Border(left=Side(border_style="+left+
        ",color="+colorL+
        "),right=Side(border_style="+right+
        ",color="+colorR+
        "),top=Side(border_style="+top+
        ",color="+colorT+
        "),bottom=Side(border_style="+bottom+
        ",color="+colorB+
        "))")
        BORD1 = str("bord"+CELL + " = " + BORD0)

        BORDRR = "Border(left=Side(border_style=None,color=None),right=Side(border_style=None,color=None),top=Side(border_style=None,color=None),bottom=Side(border_style=None,color=None))"
        if BORD0 != BORDRR:
            file1.write(BORD1)
            file1.write("\n")
            BORD2 = "ws[\""+CELL+ "\"].border = bord"+CELL
            BORD2= str(BORD2)
            file1.write(BORD2)
            file1.write("\n")



print("")
i = str(i)
print("The program scaned "+ i + " cells")
print("")
file1.write("wb.save('output.xlsx')")
file1.write("\n")

file1.write("print(\""+READY+"\")")




file1.close()
print("FORMAT.py is ready")
